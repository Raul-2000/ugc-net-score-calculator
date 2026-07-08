import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import io
import re

def extract_response_sheet(pdf_file):
    """
    Robustly extracts Question IDs and Chosen Options from the response sheet PDF.
    Tracks context sequentially line-by-line to prevent index misalignment errors.
    """
    records = []
    current_q_id = None
    
    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue
            lines = text.split("\n")
            for line in lines:
                # Standardize whitespace format around delimiters
                line_clean = line.replace(" :", ":")
                
                # Capture target Question ID block
                if "Question ID:" in line_clean:
                    parts = line_clean.split("Question ID:")
                    if len(parts) > 1:
                        current_q_id = parts[1].strip()
                
                # Pair with its matching Chosen Option contextually
                elif "Chosen Option:" in line_clean and current_q_id:
                    parts = line_clean.split("Chosen Option:")
                    chosen = parts[1].strip() if len(parts) > 1 else "--"
                    
                    records.append({
                        "Question ID": current_q_id,
                        "Chosen Option": chosen
                    })
                    current_q_id = None  # Reset pointer for the next question block
                    
    df = pd.DataFrame(records)
    if not df.empty:
        df.drop_duplicates(subset=["Question ID"], keep="last", inplace=True)
    return df

def extract_answer_key(pdf_file):
    """
    Extracts Question IDs and Correct Options from the official answer key PDF.
    """
    records = []
    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue
            lines = text.split("\n")
            for line in lines:
                # Pattern matches standard 10-digit Question IDs followed by Option IDs
                match = re.search(r'(\d{10})\s+(\d+)', line)
                if match:
                    q_id = match.group(1)
                    correct_opt = match.group(2)
                    records.append({
                        "Question ID": q_id,
                        "Correct Option": correct_opt
                    })
    df = pd.DataFrame(records)
    if not df.empty:
        df.drop_duplicates(subset=["Question ID"], inplace=True)
    return df

def calculate_score(response_pdf, answer_key_pdf):
    """
    Processes inputs, computes evaluation matrices, writes structured reports,
    and returns localized statistical parameters along with the raw document stream.
    """
    df_response = extract_response_sheet(response_pdf)
    df_key = extract_answer_key(answer_key_pdf)

    if df_response.empty or df_key.empty:
        raise ValueError("Extraction yielded empty datasets. Verify document structure accuracy.")

    # Execute a Left Join to isolate processing to official answer key structures exclusively
    combined_df = pd.merge(df_key, df_response, on="Question ID", how="left")
    
    # Standardize blank or unmatched values as explicit unattempted entries
    combined_df["Chosen Option"] = combined_df["Chosen Option"].fillna("--")

    def evaluate_row(row):
        chosen = str(row["Chosen Option"]).strip()
        correct = str(row["Correct Option"]).strip()
        if chosen in ["--", "", "-"]:
            return "Unattempted"
        elif chosen == correct:
            return "Correct"
        else:
            return "Incorrect"

    combined_df["Mark"] = combined_df.apply(evaluate_row, axis=1)
    combined_df = combined_df[["Question ID", "Chosen Option", "Correct Option", "Mark"]]

    # Calculate final examination metrics
    correct_count = int((combined_df["Mark"] == "Correct").sum())
    incorrect_count = int((combined_df["Mark"] == "Incorrect").sum())
    unattempted_count = int((combined_df["Mark"] == "Unattempted").sum())
    
    marks_gained = correct_count * 2
    marks_missed = incorrect_count * 2  # Marks missed due to wrong answers
    total_score = marks_gained

    # Build spreadsheet structure
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Evaluation Report"

    # Write target table columns
    for r in dataframe_to_rows(combined_df, index=False, header=True):
        worksheet.append(r)

    # Append clean summary block separated down below the matrix data rows
    row_offset = len(combined_df) + 3
    summary_data = [
        ("Total Questions", len(combined_df)),
        ("Correct Answers", correct_count),
        ("Incorrect Answers", incorrect_count),
        ("Unattempted Questions", unattempted_count),
        ("Marks Gained", marks_gained),
        ("Final Score Secured", total_score)
    ]
    
    for label, val in summary_data:
        worksheet.cell(row=row_offset, column=1, value=label)
        worksheet.cell(row=row_offset, column=2, value=val)
        row_offset += 1

    excel_buffer = io.BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)

    return excel_buffer.getvalue(), total_score, marks_gained, marks_missed, correct_count
